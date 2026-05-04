# -*- coding: utf-8 -*-
"""
Athens Trip April 2026 - Excel Generator
"""

import re
import os
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPTS_DIR  = os.path.dirname(__file__)
DATA_DIR     = os.path.join(SCRIPTS_DIR, "data")
OUTPUT_FILE  = os.path.join(SCRIPTS_DIR, "athens_trip_full.xlsx")

# ── צבעים ──────────────────────────────────────────────────────────────────
COLOR = {
    # לוז — צבעים לפי בקשה
    "day1":        "94DCF8",  # כחול            — יום א
    "day2":        "F7C7AC",  # אפרסק           — יום ב
    "day3":        "E49EDD",  # סגול            — יום ג
    "day4":        "7CD367",  # סגול            — יום ד
    "evening":     "ADADAD",  # אפור            — בילויי ערב
    # כותרות
    "header":      "1A252F",
    "header_font": "FFFFFF",
    # משימות
    "task_done":   "A9DFBF",
    "task_todo":   "F5CBA7",
    # עלויות
    "cost_flight": "AED6F1",
    "cost_hotel":  "A9DFBF",
    "cost_act":    "F9E79F",
    "cost_total":  "1A252F",
    "cost_total_font": "FFFFFF",
    # מסעדות
    "meat":        "F1948A",  # בשרי — אדמדם
    "dairy":       "AED6F1",  # חלבי — כחול
}

DAY_COLORS = {
    "יום א": COLOR["day1"],
    "יום ב": COLOR["day2"],
    "יום ג": COLOR["day3"],
    "יום ד": COLOR["day4"],
}

EVENING_KEYWORDS = ["ערב", "בר גג", "פלאקה בלילה", "סינטגמה", "21:00", "20:00+", "19:30+"]

# ── עזר: האם הטקסט אנגלית/ספרות בלבד? ────────────────────────────────────
def is_ltr_only(val):
    """True כאשר אין תווי עברית/ערבית — טקסט אנגלי/מספרי בלבד."""
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    return not bool(re.search(r'[\u0590-\u05FF\uFB1D-\uFB4F]', s))

def smart_align(val, wrap=True):
    """תמיד ימין."""
    return Alignment(horizontal="right", vertical="center", wrap_text=wrap, readingOrder=2)

def ltr_align(wrap=True):
    """שמאל — לעמודת קישורים."""
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap, readingOrder=1)

def right_align(wrap=True):
    return Alignment(horizontal="right", vertical="center", wrap_text=wrap, readingOrder=2)

def center_align(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap, readingOrder=2)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def header_font():
    return Font(bold=True, color=COLOR["header_font"], name="Arial", size=10)

def regular_font(bold=False, size=11):
    return Font(bold=bold, name="Arial", size=size)

def thin_border():
    s = Side(style="thin", color="666666")
    return Border(left=s, right=s, top=s, bottom=s)

def set_rtl(ws):
    ws.sheet_view.rightToLeft = True

def apply_borders_to_range(ws, min_row, max_row, min_col, max_col):
    """מחיל בורדר דק על כל התאים בטווח, כולל תאים ריקים."""
    b = thin_border()
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = b

def style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill(COLOR["header"])
        cell.font = header_font()
        cell.alignment = center_align()
        cell.border = thin_border()

def freeze_and_filter(ws, freeze_cell, filter_range):
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = filter_range

def apply_row_style(row_cells, bg_color, use_smart_align=True):
    for cell in row_cells:
        cell.fill = fill(bg_color)
        cell.font = regular_font()
        cell.border = thin_border()
        if use_smart_align:
            cell.alignment = smart_align(cell.value)
        else:
            cell.alignment = right_align()

def set_all_row_heights(ws, height=20):
    """גובה קבוע לכל השורות (מלבד כותרת)."""
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = height

def autofit_row_heights(ws, col_widths_map=None, min_h=20, line_height=13):
    """לאחר עדכון — תמיד גובה קבוע 20."""
    set_all_row_heights(ws, height=20)


# ── שיט 1: לוז יום-יום ─────────────────────────────────────────────────────
def build_itinerary(wb):
    ws = wb.create_sheet("📅 לוז")
    set_rtl(ws)

    headers = ["יום", "תאריך", "שעה", "פעילות", "מיקום/כתובת", "הסבר", "הערות", "נגיש לסבתא", "מחיר משוער"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    rows = [
        # יום א
        ("יום א", "26/04/2026", "08:45", "נחיתה בשדה התעופה", "Athens International Airport", "", "", "כן", ""),
        ("יום א", "26/04/2026", "10:30", "הגעה למלון וצ'ק-אין", "Titania Hotel Athens", "", "אם החדרים לא מוכנים — השאירו מזוודות", "כן", ""),
        ("יום א", "26/04/2026", "11:00-13:00", "הופ-און הופ-אוף", "מרכז אתונה", "אוטובוס תיירותי קומתיים פתוח. קונים כרטיס 48 שעות ועולים/יורדים בכל תחנה. מקבלים אוזניות שם.", "מומלץ ביום ראשון — עייפים!", "כן — ישיבה", "~28€ מבוגר / ~14€ ילד"),
        ("יום א", "26/04/2026", "13:30", "צהריים — Shuk (בשרי)", "Ermou 117", "", "הזמינו מקום ל-12!", "כן", "~20-25€/אדם"),
        ("יום א", "26/04/2026", "14:45-17:45", "סיור Ezraider", "Leof. Andrea Siggrou 22 — GoGo Electric", "כלי רכב חשמלי ישראלי על 4 גלגלים. מדריך עברית, עוצרים באטרקציות.", "🚖 לקחת אובר! לאשר יציאה ב-14:45 + נגישות לסבתא", "לאשר", "59.50€ מבוגר / 29.75€ ילד"),
        ("יום א", "26/04/2026", "18:00-19:00", "Who Killed Callimachos — מיסתורן", "Athens Living Museum", "מופע תיאטרון אינטראקטיבי. שחקן מגלם דמויות מאתונה העתיקה, הקהל הם הבלשים.", "הנחה ל-10+ אנשים! הזמינו מראש.", "כן", "22€ מבוגר / 15€ ילד"),
        ("יום א", "26/04/2026", "19:00-20:00", "מנוחה במלון", "Titania Hotel Athens", "", "חובה אחרי טיסת שחר", "כן", ""),
        ("יום א", "26/04/2026", "20:00+", "ערב בכיכר סינטגמה", "Syntagma Square", "כיכר המרכזית של אתונה — מוארת ותוססת. בתי קפה, ברים, אנשים.", "שתייה מבקבוקים סגורים בסדר", "כן", ""),
        # יום ב
        ("יום ב", "27/04/2026", "09:30-11:30", "שכונת פלאקה", "Plaka neighborhood", "שכונת העיר העתיקה. רחובות מרוצפים מהמאה ה-19, חנויות מזכרות, אווירה יוונית.", "שטוח ונגיש לסבתא", "כן", "כניסה חופשית"),
        ("יום ב", "27/04/2026", "11:30-13:30", "קניות", "Ermou + Monastiraki", "רחוב ארמו — קניות ראשי. מונסטיראקי — מזכרות.", "קחו תיק לקניות", "כן — שטוח", "—"),
        ("יום ב", "27/04/2026", "13:30", "צהריים — Gostijo (בשרי)", "Esopou 10", "ממש ליד מונסטיראקי. פותח ב-13:30. יש מכולת כשרה בפנים.", "לאשר זמינות ל-12", "כן", "~20-25€/אדם"),
        ("יום ב", "27/04/2026", "14:30-18:00", "המשך קניות", "Ermou + Monastiraki", "המשך קניות בארמו ומונסטיראקי.", "", "כן — שטוח", "—"),
        ("יום ב", "27/04/2026", "18:30-20:30", "גבעת ליקבטוס + שקיעה", "Lycabettus Hill", "הגבעה הכי גבוהה באתונה (277מ'). נוף פנורמי. שקיעה ~20:15 — לא לפספס!", "רכבל תת-קרקעי — 3 דקות. להגיע ב-18:30!", "כן — רכבל", "13€ הלוך-חזור"),
        ("יום ב", "27/04/2026", "21:00+", "בר גג לשתייה", "Rooftop bar", "נוף לילי על העיר המוארת.", "שתייה מבקבוקים סגורים בסדר", "לשיקול דעת", ""),
        # יום ג
        ("יום ג", "28/04/2026", "10:00", "יציאה לנפפליאו — ואן פרטי", "Titania Hotel Athens", "קחו אוכל כשר מהמלון — אין מסעדה כשרה בנפפליאו!", "", "כן — ואן", "650€ לקבוצה (~54€/אדם)"),
        ("יום ג", "28/04/2026", "11:00", "תעלת קורינטוס", "Corinth Canal", "תעלה צרה ועמוקה מרהיבה. הישג הנדסי מדהים. עצירה ~20-30 דקות.", "נוף מדהים מהגשר. שטוח ונגיש.", "כן", "חינם"),
        ("יום ג", "28/04/2026", "12:30-13:00", "הגעה לנפפליאו", "Nafplio", "", "", "כן — ואן", ""),
        ("יום ג", "28/04/2026", "13:00-14:30", "מבצר פלמידי", "Fortress Palamidi", "אחד המבצרים היפים ביוון. נוף פנורמי על העיירה והמפרץ.", "סבתא יכולה לחכות בנמל — הרבה מדרגות!", "חלקי", "~8€/אדם"),
        ("יום ג", "28/04/2026", "14:30-16:30", "טיול חופשי בנפפליאו", "Nafplio — Old Town + Harbor", "סמטאות ציוריות, ארכיטקטורה וונציאנית, נמל יפהפה.", "שטוח בעיירה ובנמל. נגיש לסבתא.", "כן — שטוח", "כניסה חופשית"),
        ("יום ג", "28/04/2026", "16:30", "סירה לבורג'י (אופציונלי)", "Nafplio Harbor", "מגדל ים באמצע המפרץ — ~15 דקות הפלגה.", "", "כן", "~5€/אדם"),
        ("יום ג", "28/04/2026", "17:00", "יציאה חזרה לאתונה", "Nafplio - Athens", "", "", "כן — ואן", ""),
        ("יום ג", "28/04/2026", "18:30", "הגעה למלון", "Titania Hotel Athens", "", "", "כן", ""),
        ("יום ג", "28/04/2026", "21:00+", "ערב בפלאקה", "Plaka neighborhood", "פלאקה יפה בלילה. אווירה רומנטית, בתי קפה שקטים.", "שתייה בסדר. ערב נינוח.", "כן", ""),
        # יום ד
        ("יום ד", "29/04/2026", "09:00-11:00", "סיור גרפיטי + פסירי", "Psiri / Kerameikos", "שכונת האמנות של אתונה. קירות מלאי מורלים. ממשיכים ישר לשוק הפשפשים!", "אפליקציית Street Art Cities (חינם)", "כן — שטוח", "חינם"),
        ("יום ד", "29/04/2026", "11:00-13:00", "מונסטיראקי ושוק הפשפשים", "Monastiraki Square", "שוק פשפשים ססגוני. תכשיטים, ביגוד, עתיקות, מזכרות.", "שטוח, הרבה ספסלים לסבתא", "כן — שטוח", "כניסה חופשית"),
        ("יום ד", "29/04/2026", "13:00-14:00", "מוזיאון האשליות", "Museum of Illusions Athens", "אשליות אופטיות ותמונות אינטראקטיביות. ~60 דקות. כיפי לכל הגילאים!", "לשאול על הנחות קבוצה ל-12", "כן", "~13€/אדם"),
        ("יום ד", "29/04/2026", "14:15", "צהריים — Shuk (בשרי)", "Ermou 117", "5 דקות הליכה מהמוזיאון", "", "כן", "~20-25€/אדם"),
        ("יום ד", "29/04/2026", "15:30-16:30", "הגן הלאומי", "National Garden", "פארק ירוק ענק ליד סינטגמה. שבילים, בריכות, צל, שקט.", "הליכה קצרה מהמלון", "כן — שטוח", "חינם"),
        ("יום ד", "29/04/2026", "16:30", "חזרה למלון — איסוף מזוודות", "Titania Hotel Athens", "", "מזוודות בקבלה מאז 12:00", "כן", ""),
        ("יום ד", "29/04/2026", "17:00", "נסיעה לשדה התעופה", "Athens International Airport", "", "מיניבוס (יאנה) או מטרו (10 euro/person). לא לאחר!", "כן — מיניבוס", "~10€/אדם"),
        ("יום ד", "29/04/2026", "18:00", "להיות בשדה התעופה!", "Athens International Airport", "", "לא לאחר!", "כן", ""),
        ("יום ד", "29/04/2026", "20:30", "טיסה לישראל", "Athens - Israel", "", "", "כן", ""),
    ]

    for row_data in rows:
        ws.append(list(row_data))

    # עיצוב שורות
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        day_val = row[0].value or ""
        time_val = row[2].value or ""
        activity_val = row[3].value or ""

        is_evening = any(k in str(time_val) + str(activity_val) for k in EVENING_KEYWORDS)
        row_color = COLOR["evening"] if is_evening else DAY_COLORS.get(day_val, "FFFFFF")

        for cell in row:
            cell.fill = fill(row_color)
            cell.font = regular_font()
            cell.border = thin_border()
            cell.alignment = smart_align(cell.value)

    # יום | תאריך | שעה | פעילות | מיקום | הסבר | הערות | נגיש | מחיר
    col_widths = [7, 12, 12.5, 34, 34, 83, 41, 12, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20
    freeze_and_filter(ws, "A2", f"A1:{get_column_letter(len(headers))}1")

    # יום | תאריך | שעה | פעילות | מיקום | הסבר | הערות | נגיש | מחיר
    final_col_widths = [7, 11, 12.5, 34, 34, 83, 41, 12, 25]
    for i, w in enumerate(final_col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    autofit_row_heights(ws)
    apply_borders_to_range(ws, 1, ws.max_row, 1, len(final_col_widths))


# ── שיט 2: משימות ──────────────────────────────────────────────────────────
def build_tasks(wb):
    ws = wb.create_sheet("✅ משימות")
    set_rtl(ws)

    headers = ["משימה", "מבוצע", "פרטי קשר / קישור", "הערות"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    tasks = [
        ("הזמנת מיניבוס ליאנה — הגעה מהשדה (26/4)", "", "https://wa.me/message/NLNHNSS46SLHE1", "12 איש + מזוודות. נסיעה ~45 דק"),
        ("הזמנת מיניבוס ליאנה — חזרה לשדה (29/4)", "", "https://wa.me/message/NLNHNSS46SLHE1", "להיות בשדה ב-18:00"),
        ("הזמנת טיול יום לנפפליאו (28/4 10:00)", "", "https://wa.me/message/NLNHNSS46SLHE1", "650 euro לקבוצה. ואן פרטי ל-12. לאשר תעלת קורינטוס + מבצר פלמידי"),
        ("לברר הזמנה מוקדמת — Shuk / King David / Gostijo", "", "Shuk: +306970252857 | King David: https://wa.me/306949527534", "כנראה לא חובה — אבל כדאי לבדוק"),
        ("הזמנת סיור Ezraider (26/4 14:45)", "", "https://wa.me/message/NLNHNSS46SLHE1", "59.50 euro מבוגר / 29.75 euro ילד (15% הנחה). לאשר יציאה ב-14:45 + נגישות לסבתא"),
        ("הזמנת Who Killed Callimachos (26/4 18:00)", "", "https://athenslivingmuseum.com/who-killed-callimachos/", "22 euro מבוגר / 15 euro ילד. הנחה ל-10+ אנשים!"),
        ("הזמנת כרטיסי הופ-און הופ-אוף אונליין", "", "https://www.bigbustours.com/en/athens/athens-bus-tours", "כרטיס 48 שעות"),
        ("אישור עם המלון — מעלית + נגישות לחדר", "", "Titania Hotel Athens - Diomeias 5", ""),
    ]

    for task in tasks:
        ws.append(list(task))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        done = str(row[1].value or "").strip().lower() in ["כן", "v", "yes"]
        row_color = COLOR["task_done"] if done else COLOR["task_todo"]
        for cell in row:
            cell.fill = fill(row_color)
            cell.font = regular_font()
            cell.border = thin_border()
            # עמודה 3 = פרטי קשר/קישור — יישור שמאל
            if cell.column == 3:
                cell.alignment = ltr_align()
            else:
                cell.alignment = right_align()

    col_widths = [45, 10, 50, 38]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20
    autofit_row_heights(ws)
    apply_borders_to_range(ws, 1, ws.max_row, 1, len(col_widths))
    freeze_and_filter(ws, "A2", f"A1:{get_column_letter(len(headers))}1")


# ── שיט 3: עלויות ──────────────────────────────────────────────────────────
def build_costs(wb):
    ws = wb.create_sheet("💰 עלויות")
    set_rtl(ws)

    headers = ["פריט", "עלות מבוגר (euro)", "עלות ילד (euro)", "עלות מבוגר (ILS)", "עלות ילד (ILS)", "כמות מבוגרים", "כמות ילדים", "סה\"כ משוער (euro)", "הערות"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    sections = {
        "flight": [
            ("טיסה — בסיס שווה לכולם", 131.59, 131.59, 471, 471, 8, 4, 1579.08, "חלוקה שווה אחרי הורדת תוספות אישיות"),
            ("תוספת — Hila (SKY enjoy)", "+82", "", "+294", "", 1, "", "", "מטען 23kg + Extra legroom"),
            ("תוספת — Nava (SKY enjoy)", "+82", "", "+294", "", 1, "", "", "מטען 23kg + Extra legroom"),
            ("תוספת — Or (SKY joy+)", "+64", "", "+229", "", 1, "", "", "מטען 15kg + שינוי חינם"),
            ("תוספת — Yehuda (SKY joy+)", "+64", "", "+229", "", 1, "", "", "מטען 15kg + שינוי חינם"),
            ("תוספת — Adir (מושב שמור)", "+14", "", "+50", "", 1, "", "", "מושב חלון 15F"),
            ("תוספת — Kfir (מושב שמור)", "+14", "", "+50", "", 1, "", "", "מושב אמצע 15E"),
            ("סה\"כ טיסות כולל תוספות", "", "", "", "", 12, "", 1899.08, ""),
        ],
        "hotel": [
            ("מלון — הזמנה 1 (59110948) 5 חדרים", 172.50, 172.50, 612, 612, 10, "", 1725, "כולל ארוחת בוקר. Climate Tax 150€ בצ'ק-אין"),
            ("מלון — הזמנה 2 (59110949) 1 חדר", 115.00, 115.00, 408, 408, 3, "", 345, "אדיר+הילה+כפיר. Climate Tax 30€ בצ'ק-אין"),
        ],
        "activities": [
            ("הסעה שדה + מלון (x2)", "~23", "~23", "~82", "~82", 13, "", "~300", "150€ לכיוון × 2 ÷ 13"),
            ("הופ-און הופ-אוף", "~28", "~14", "~100", "~50", 8, 4, "~280", "כרטיס 48 שעות"),
            ("סיור Ezraider", 59.50, 29.75, 213, 106, 9, 4, "~655", "15% הנחה. מחיר מקורי 70/35€"),
            ("Who Killed Callimachos", 22, 15, 79, 54, 8, 4, "~236", "הנחה קבוצתית ל-10+"),
            ("טיול יום — נפפליאו (ואן פרטי)", "~54", "~54", "~193", "~193", 8, 4, 650, "650 euro / 12. כולל תעלת קורינטוס"),
            ("כניסה מבצר פלמידי", "~8", "~8", "~29", "~29", 8, 4, "~96", "ילדים עד 18 — חינם"),
            ("גבעת ליקבטוס (רכבל)", 13, 13, 47, 47, 8, 4, "~156", "הלוך-חזור"),
            ("מוזיאון האשליות", "~13", "~13", "~47", "~47", 8, 4, "~156", "הנחות קבוצה — לשאול"),
            ("ארוחות צהריים x4", "~22", "~15", "~79", "~54", 8, 4, "~816", "ממוצע"),
            ("קניות / קפה / שונות", "~30", "~30", "~107", "~107", 12, "", "~360", "לפי שיקול דעת אישי"),
        ],
        "totals": [
            ("סה\"כ פעילויות+אוכל — מבוגר", "~270-320", "", "~967-1146", "", "", "", "", "לא כולל קניות אישיות"),
            ("סה\"כ פעילויות+אוכל — ילד/נוער", "~193-233", "", "~691-834", "", "", "", "", "הנחות כניסה / פחות אוכל"),
            ("סה\"כ כולל הכל — מבוגר (בסיס)", "~599", "", "~2,144", "", "", "", "", "שער 3.58 ILS/euro (15/4/2026)"),
            ("סה\"כ כולל הכל — ילד/נוער", "~517", "", "~1,851", "", "", "", "", "שער 3.58 ILS/euro (15/4/2026)"),
        ]
    }

    section_colors = {
        "flight":     COLOR["cost_flight"],
        "hotel":      COLOR["cost_hotel"],
        "activities": COLOR["cost_act"],
        "totals":     COLOR["cost_total"],
    }

    for section, section_rows in sections.items():
        for row_data in section_rows:
            ws.append(list(row_data))
            row_idx = ws.max_row
            is_total = section == "totals"
            row_color = section_colors[section]
            for cell in ws[row_idx]:
                cell.fill = fill(row_color)
                if is_total:
                    cell.font = Font(bold=True, color=COLOR["cost_total_font"], name="Arial", size=9)
                else:
                    cell.font = regular_font()
                cell.border = thin_border()
                cell.alignment = smart_align(cell.value)

    col_widths = [35, 16, 14, 16, 14, 14, 12, 18, 38]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20
    autofit_row_heights(ws)
    apply_borders_to_range(ws, 1, ws.max_row, 1, len(col_widths))
    ws.freeze_panes = "A2"


# ── שיט 4: נוסעים ──────────────────────────────────────────────────────────
def build_passengers(wb):
    ws = wb.create_sheet("👥 נוסעים")
    set_rtl(ws)

    headers = ["שם", "הזמנה", "כרטיס", "כבודה", "בסיס (euro)", "תוספת (euro)", "סה\"כ טיסה (euro)", "מלון (euro)", "סה\"כ טיסה+מלון (euro)"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    passengers = [
        ("Adi Huri",    "0L4ZFG", "SKY joy",   "יד 8kg",                       131.59, "—",   131.59, 172.50, 304.09),
        ("Shira Huri",  "0L4ZFG", "SKY joy",   "יד 8kg",                       131.59, "—",   131.59, 172.50, 304.09),
        ("Yovel Zvigi", "0L4ZFG", "SKY joy",   "יד 8kg",                       131.59, "—",   131.59, 172.50, 304.09),
        ("Agam Zvigi",  "FW9ZS7", "SKY joy",   "יד 8kg",                       131.59, "—",   131.59, 172.50, 304.09),
        ("Hana Huri",   "FW9ZS7", "SKY joy",   "יד 8kg",                       131.59, "—",   131.59, 172.50, 304.09),
        ("Hila Huri",   "GEIMP8", "SKY enjoy", "יד + מטען 23kg, Extra legroom", 131.59, "+82", 213.59, 115.00, 328.59),
        ("Noam Huri",   "UDJWVK", "SKY joy",   "יד 8kg",                       131.59, "—",   131.59, 172.50, 304.09),
        ("Or Zvigi",    "V2FDEL", "SKY joy+",  "יד + מטען 15kg",               131.59, "+64", 195.59, 172.50, 368.09),
        ("Yehuda Huri", "V2FDEL", "SKY joy+",  "יד + מטען 15kg",               131.59, "+64", 195.59, 172.50, 368.09),
        ("Adir Huri",   "WR7O1Y", "SKY joy",   "יד 8kg, מושב 15F",             131.59, "+14", 145.59, 115.00, 260.59),
        ("Kfir Huri",   "WR7O1Y", "SKY joy",   "יד 8kg, מושב 15E",             131.59, "+14", 145.59, 115.00, 260.59),
        ("Nava Huri",   "0GI2VD", "SKY enjoy", "יד + מטען 23kg, Extra legroom", 131.59, "+82", 213.59, 172.50, 386.09),
        ("Yarin Zvigi", "JN915E", "SKY joy",   "יד 8kg",                       131.59, "—",   131.59, 172.50, 304.09),
    ]

    row_colors = [COLOR["day1"], "D6EAF8"]
    for i, p in enumerate(passengers):
        ws.append(list(p))
        row_idx = ws.max_row
        c = row_colors[i % 2]
        for cell in ws[row_idx]:
            cell.fill = fill(c)
            cell.font = regular_font()
            cell.border = thin_border()
            cell.alignment = smart_align(cell.value)

    col_widths = [18, 12, 12, 30, 14, 14, 18, 14, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20
    autofit_row_heights(ws)
    apply_borders_to_range(ws, 1, ws.max_row, 1, len(col_widths))
    ws.freeze_panes = "A2"


# ── שיט 5: מסעדות ──────────────────────────────────────────────────────────
def build_restaurants(wb):
    ws = wb.create_sheet("🍽️ מסעדות")
    set_rtl(ws)

    headers = ["סוג", "שם", "כתובת", "שעות פתיחה", "כשרות", "מה יש שם", "הערות", "איפה בלוז"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    restaurants = [
        # בשרי
        ("בשרי", "Shuk (השוק)", "Ermou 117", "09:00-22:30", "בד\"צ למהדרין", "פלאפל, שקשוקה, חומוס, שניצל, דגים", "הכי מומלצת. קרובה למלון.", "יום א צהריים + יום ד צהריים"),
        ("בשרי", "King David Burger", "Ermou 78", "א-ה 11:00-23:00\nמוצ\"ש עד 01:00", "בית יוסף", "המבורגרים, מעורב ירושלמי, עראיס. גם משלוחים בוולט.", "", ""),
        ("בשרי", "Gostijo / חב\"ד", "Esopou 10", "א-ה 13:30-21:30", "בית חב\"ד", "אוכל בית חב\"ד. יש מכולת כשרה בפנים!", "מאובטח, צריך דרכון.", "יום ב צהריים"),
        # חלבי
        ("חלבי", "Parakalo", "Mikonos 18", "א-ה 8:00-20:00\nו עד 15:00", "חב\"ד", "ארוחות בוקר, פיצות, מאפים, סלטים.", "מומלץ לארוחות בוקר!", "יום ב/ג/ד בוקר"),
    ]

    for rest in restaurants:
        ws.append(list(rest))
        r = ws.max_row
        is_dairy = rest[0] == "חלבי"
        row_color = COLOR["dairy"] if is_dairy else COLOR["meat"]
        for cell in ws[r]:
            cell.fill = fill(row_color)
            cell.font = regular_font()
            cell.border = thin_border()
            cell.alignment = smart_align(cell.value)
        ws.row_dimensions[r].height = 45

    # מקרא — שורה ריקה ואז בשרי/חלבי
    ws.append([])
    empty_row = ws.max_row
    for cell in ws[empty_row]:
        cell.fill = fill("FFFFFF")

    ws.append(["בשרי", "", "", "", "", "", "", ""])
    legend_row = ws.max_row
    for cell in ws[legend_row]:
        cell.fill = fill(COLOR["meat"])
        cell.font = regular_font(bold=True)
        cell.border = thin_border()
        cell.alignment = right_align()

    ws.append(["חלבי", "", "", "", "", "", "", ""])
    for cell in ws[ws.max_row]:
        cell.fill = fill(COLOR["dairy"])
        cell.font = regular_font(bold=True)
        cell.border = thin_border()
        cell.alignment = right_align()

    col_widths = [10, 20, 18, 22, 15, 45, 30, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20
    autofit_row_heights(ws)
    apply_borders_to_range(ws, 1, ws.max_row, 1, len(col_widths))
    freeze_and_filter(ws, "A2", f"A1:{get_column_letter(len(headers))}1")


# ── שיט 6: אטרקציות ────────────────────────────────────────────────────────
def build_attractions(wb):
    ws = wb.create_sheet("🗺️ אטרקציות")
    set_rtl(ws)

    def write_section_title(row, col, text):
        cell = ws.cell(row=row, column=col)
        cell.value = text
        cell.font = Font(bold=True, name="Arial", size=12, color="1A252F")
        cell.alignment = right_align()

    def write_header_row(row, start_col, headers, bg="34495E"):
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=start_col + i)
            cell.value = h
            cell.fill = fill(bg)
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            cell.alignment = center_align()
            cell.border = thin_border()

    # ── חלק א: אקרופוליס (אופציונלי) ──────────────────────────────────────
    write_section_title(1, 1, "האקרופוליס — מידע כללי (אופציונלי)")
    write_header_row(2, 1, ["אטרקציה", "שם", "תיאור", "הערות מעשיות"], bg="566573")

    acrop_data = [
        ("האקרופוליס", "Acropolis Hill",
         "גובה סלע 156 מטר שעליו עתיקים בני 2,500 שנה — בראשם הפרתנון, מקדש אתנה. הסמל של יוון ואחד מהמורשות החשובים בעולם.",
         "קן כרטיסים אונליין מראש! (נסיעה + מעלית פנורמית). תאמו 24ש מראש: +30 4172 321 210"),
        ("מוזיאון האקרופוליס", "Acropolis Museum",
         "מוזיאון הבנוי מתחת לאקרופוליס, מציג את האפסלות והממצאים המקוריים מהגבעה. רצפת זכוכית עם חפירות ארכיאולוגיות מתחתיה.",
         "נגיש לחלוטין — מעלית ורמפות. מדהים!"),
    ]

    for i, row_data in enumerate(acrop_data):
        r = 3 + i
        for j, val in enumerate(row_data):
            cell = ws.cell(row=r, column=1 + j)
            cell.value = val
            cell.fill = fill("DAE9F8")
            cell.font = regular_font()
            cell.border = thin_border()
            cell.alignment = smart_align(val)
        ws.row_dimensions[r].height = 20

    # ── חלק ב: אטרקציות כלליות ─────────────────────────────────────────────
    write_section_title(6, 1, "אטרקציות — מידע כללי")
    write_header_row(7, 1, ["קטגוריה", "שם האטרקציה", "תיאור", "הערות מעשיות", "מחיר משוער", "נגישות"])

    attractions = [
        ("מוזיאון",    "מוזיאון האשליות",         "אשליות אופטיות ותמונות אינטראקטיביות. ~60 דקות. כיפי לכל הגילאים.",      "פתוח מ-10:00. הנחות קבוצה ל-12.",   "~13€/אדם",              "כן"),
        ("תיאטרון",    "Who Killed Callimachos",  "מיסתורן אינטראקטיבי. שחקן + קהל בלשים. אתונה עתיקה.",                   "מדי יום 18:00. הנחה ל-10+.",         "22€ מבוגר / 15€ ילד",  "כן"),
        ("סיור",       "Ezraider",                "כלי חשמלי ל-4 גלגלים. מדריך עברית. היסטוריה יהודית.",                   "🚖 אובר ל-Leof. Andrea Siggrou 22. לאשר ב-14:45.", "59.50€ מבוגר / 29.75€ ילד (15% הנחה)",  "לאשר"),
        ("טיול יום",  "נפפליאו + תעלת קורינטוס", "עיירת בוטיק ציורית + תעלה הנדסית. ואן פרטי.",                            "10:00-18:30. אוכל מהבית!",           "650€ לקבוצה (~54€)",   "כן — ואן"),
        ("מבצר",       "מבצר פלמידי",             "אחד המבצרים היפים ביוון. נוף פנורמי.",                                   "הרבה מדרגות — סבתא נשארת בנמל.",   "~8€/אדם",              "חלקי"),
        ("גבעה",       "גבעת ליקבטוס",            "הגבעה הכי גבוהה. נוף 360 מעלות. שקיעה מרהיבה ~20:15.",                 "רכבל עד 02:30. להגיע ב-18:30.",     "13€ הלוך-חזור",        "כן — רכבל"),
        ("פארק",       "הגן הלאומי",              "פארק ירוק ענק ליד סינטגמה. שבילים, בריכות, צל.",                        "פתוח עד שקיעה.",                     "חינם",                 "כן — שטוח"),
        ("שכונה",      "פלאקה",                   "שכונת העיר העתיקה. רחובות מרוצפים, מזכרות, קפה.",                       "שטוח ונגיש. יפה גם בלילה.",          "כניסה חופשית",         "כן"),
        ("שכונה",      "פסירי",                   "שכונת הגרפיטי של אתונה. מורלים, אמנות רחוב.",                           "אפליקציה: Street Art Cities.",       "חינם",                 "כן — שטוח"),
        ("שוק",        "מונסטיראקי",              "כיכר + שוק פשפשים ססגוני. תכשיטים, ביגוד, עתיקות.",                     "שטוח. הרבה ספסלים לסבתא.",          "כניסה חופשית",         "כן"),
        ("תחבורה",     "הופ-און הופ-אוף",         "אוטובוס קומתי פתוח. עולים/יורדים בכל תחנה. כרטיס 48 שעות.",            "bigbustours.com — לרכוש מראש.",     "~28€ מבוגר / ~14€ ילד","כן — ישיבה"),
        ("טיול ים",    "סירה לבורג'י",            "מגדל ים באמצע מפרץ נפפליאו. ~15 דקות הפלגה.",                           "אופציונלי. מהנמל.",                  "~5€/אדם",              "כן"),
    ]

    row_colors = ["EBF5FB", "FDFEFE"]
    for i, attr in enumerate(attractions):
        r = 8 + i
        for j, val in enumerate(attr):
            cell = ws.cell(row=r, column=1 + j)
            cell.value = val
            cell.fill = fill(row_colors[i % 2])
            cell.font = regular_font()
            cell.border = thin_border()
            cell.alignment = smart_align(val)
        ws.row_dimensions[r].height = 20

    # רוחב עמודות
    col_widths = [14, 28, 60, 38, 22, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[6].height = 20
    autofit_row_heights(ws)
    # אקרופוליס נשאר 20 — כבר הוגדר, autofit שומר
    apply_borders_to_range(ws, 1, ws.max_row, 1, len(col_widths))
    ws.freeze_panes = "A2"


# ── ייצור הקובץ ─────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_itinerary(wb)
    build_tasks(wb)
    build_costs(wb)
    build_passengers(wb)
    build_restaurants(wb)
    build_attractions(wb)

    wb.save(OUTPUT_FILE)
    print(f"OK - Excel file created: {OUTPUT_FILE}")

    # העתק לתיקיית downloads של האתר (site/scripts/ → site/public/downloads/)
    downloads_dir = os.path.join(os.path.dirname(__file__), "..", "public", "downloads")
    os.makedirs(downloads_dir, exist_ok=True)
    import shutil
    dest = os.path.join(downloads_dir, "athens_trip_full.xlsx")
    shutil.copy2(OUTPUT_FILE, dest)
    print(f"OK - Copied to: {dest}")


if __name__ == "__main__":
    main()
