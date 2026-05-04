"""
חישוב סופי של חלוקת הוצאות — אתונה 2026
מפיק:
  - athens_settlement.md — סיכום ב-Markdown
  - athens_settlement.xlsx — Excel עם שני גיליונות
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── רשימת אנשים ────────────────────────────────────────────────────────
PEOPLE = [
    "יהודה", "נאוה", "אדיר", "הילה", "כפיר", "עדי", "שירה",
    "נועם", "אגם", "יובל", "חנה", "אור", "ירין",
]

# ─── הוצאות ────────────────────────────────────────────────────────────
expenses = [
    {
        "desc": "מונית חזור לשדה",
        "total": 150.00,
        "paid_by": {"יהודה": 150.00},
        "split_among": PEOPLE,
    },
    {
        "desc": "טיול 28/4 (כולל טיפ)",
        "total": 675.00,
        "paid_by": {"יהודה": 200.00, "אדיר": 350.00, "נאוה": 125.00},
        "split_among": PEOPLE,
    },
    {
        "desc": "רכבל גבעה (ליקבטוס)",
        "total": 91.00,
        "paid_by": {"יהודה": 91.00},
        "split_among": ["יהודה", "חנה", "נאוה", "אור", "אגם", "יובל", "ירין"],
    },
    {
        "desc": "מוזיאון האשליות — מבוגרים",
        "total": 137.50,
        "paid_by": {"הילה": 137.50},
        "split_among": [
            "יהודה", "אדיר", "הילה", "עדי", "שירה",
            "אגם", "יובל", "חנה", "אור", "ירין", "נועם",
        ],
    },
    {
        "desc": "מוזיאון האשליות — כפיר (ילד)",
        "total": 9.50,
        "paid_by": {"הילה": 9.50},
        "split_among": ["כפיר"],
    },
    {
        "desc": "Climate Tax (מלון)",
        "total": 180.00,
        "paid_by": {"אגם": 30.00, "הילה": 60.00, "נאוה": 30.00, "יובל": 60.00},
        "split_among": PEOPLE,
    },
    {
        "desc": "איזיריידר",
        "total": 655.00,
        "paid_by": {"יהודה": 655.00},
        "split_among": None,
        "custom_split": {
            "אדיר": 44.625, "כפיר": 44.625,
            "חנה": 44.625, "אור": 44.625,
            "אגם": 44.625, "יובל": 44.625,
            "נאוה": 44.625, "ירין": 44.625,
            "הילה": 59.50, "עדי": 59.50, "נועם": 59.50, "שירה": 59.50,
            "יהודה": 59.50,
        },
    },
    {
        "desc": "הסעה מהשדה",
        "total": 150.00,
        "paid_by": {"אדיר": 150.00},
        "split_among": PEOPLE,
    },
    {
        "desc": "מלון Titania (5 חדרים)",
        "total": 1575.00,
        "paid_by": {"יהודה": 1575.00},
        "split_among": [
            "יהודה", "נאוה", "עדי", "שירה", "נועם",
            "אגם", "יובל", "חנה", "אור", "ירין",
        ],
    },
]

# ─── חישוב חיוב לכל אדם ────────────────────────────────────────────────
owes = defaultdict(float)
paid = defaultdict(float)

for exp in expenses:
    for person, amount in exp["paid_by"].items():
        paid[person] += amount

    if exp.get("custom_split"):
        for person, amount in exp["custom_split"].items():
            owes[person] += amount
    elif exp["split_among"]:
        per_person = exp["total"] / len(exp["split_among"])
        for person in exp["split_among"]:
            owes[person] += per_person

balances = {p: round(paid[p] - owes[p], 2) for p in PEOPLE}

# שער ההמרה: יורו → שקלים
EUR_TO_ILS = 3.5
def ils(eur: float) -> float:
    return round(eur * EUR_TO_ILS, 2)

# ─── איחוד משפחת חורי גדולה (אדיר משלם בשבילם) ─────────────────────────
HURI_FAMILY = ["אדיר", "הילה", "כפיר", "עדי", "שירה", "נועם"]
INDIVIDUALS = ["יהודה", "נאוה", "אגם", "יובל", "חנה", "אור", "ירין"]

# יוצרים מבנה balances חלופי: אדיר מייצג את כל המשפחה
balances_family = {}
paid_family = defaultdict(float)
owes_family = defaultdict(float)

# אדיר = סכום של 6 בני המשפחה
balances_family["משפחת חורי (אדיר)"] = sum(balances[p] for p in HURI_FAMILY)
paid_family["משפחת חורי (אדיר)"] = sum(paid[p] for p in HURI_FAMILY)
owes_family["משפחת חורי (אדיר)"] = sum(owes[p] for p in HURI_FAMILY)

# שאר היחידים — נשארים כמו שהם
for p in INDIVIDUALS:
    balances_family[p] = balances[p]
    paid_family[p] = paid[p]
    owes_family[p] = owes[p]

balances_family = {k: round(v, 2) for k, v in balances_family.items()}

# ─── אימות ─────────────────────────────────────────────────────────────
total_paid = sum(paid.values())
total_owes = sum(owes.values())
print(f"סה\"כ שולם: EUR {total_paid:.2f}")
print(f"סה\"כ חייב: EUR {total_owes:.2f}")
print(f"הפרש: EUR {total_paid - total_owes:.2f}")
print(f"סך כל היתרות: EUR {sum(balances.values()):.2f}\n")

# ─── Debt Simplification ───────────────────────────────────────────────
def simplify_debts(balances_dict):
    bals = {p: round(b, 2) for p, b in balances_dict.items()}
    creditors = [(p, b) for p, b in bals.items() if b > 0.01]
    debtors = [(p, -b) for p, b in bals.items() if b < -0.01]
    creditors.sort(key=lambda x: -x[1])
    debtors.sort(key=lambda x: -x[1])

    transfers = []
    i, j = 0, 0
    while i < len(debtors) and j < len(creditors):
        debtor_name, debt = debtors[i]
        creditor_name, credit = creditors[j]
        amount = min(debt, credit)
        if amount > 0.01:
            transfers.append((debtor_name, creditor_name, round(amount, 2)))
        debtors[i] = (debtor_name, debt - amount)
        creditors[j] = (creditor_name, credit - amount)
        if debtors[i][1] < 0.01:
            i += 1
        if creditors[j][1] < 0.01:
            j += 1
    return transfers

transfers = simplify_debts(balances)
transfers_family = simplify_debts(balances_family)

# ─── הדפסה ──────────────────────────────────────────────────────────────
print("=== יתרה לכל אדם ===")
for p in PEOPLE:
    sign = "+" if balances[p] >= 0 else ""
    print(f"  {p:8s}  שילם EUR {paid[p]:8.2f}  חייב EUR {owes[p]:8.2f}  יתרה {sign}EUR {balances[p]:8.2f}")

print("\n=== העברות סופיות ===")
for src, dst, amt in transfers:
    print(f"  {src} -> {dst}: EUR {amt:.2f}")

# ═══════════════════════════════════════════════════════════════════════
# Markdown
# ═══════════════════════════════════════════════════════════════════════
md_lines = [
    "# 💳 סיכום הוצאות אתונה 2026",
    "",
    f'**סה"כ הוצאות:** €{total_paid:,.2f}  |  **משתתפים:** 13 אנשים',
    "",
    "---",
    "",
    "## 📋 פירוט הוצאות",
    "",
    "| הוצאה | סכום (€) | סכום (₪) | מי שילם | מי משתתף |",
    "|------|------|------|---------|----------|",
]
for exp in expenses:
    paid_str = ", ".join(f"{p} €{a:.2f}" for p, a in exp["paid_by"].items())
    if exp.get("custom_split"):
        participants_str = "13 (זוגות €44.625 / בודדים €59.50)"
    elif exp["split_among"]:
        participants_str = f"{len(exp['split_among'])} - {', '.join(exp['split_among'])}"
    else:
        participants_str = "—"
    md_lines.append(f"| {exp['desc']} | €{exp['total']:.2f} | ₪{ils(exp['total']):,.2f} | {paid_str} | {participants_str} |")

md_lines += [
    "",
    "---",
    "",
    "## 💰 יתרה לכל אדם",
    "",
    "| אדם | שילם | חייב | יתרה (€) | יתרה (₪) |",
    "|-----|------|------|------|------|",
]
for p in PEOPLE:
    bal = balances[p]
    if bal > 0.01:
        sign = "🟢 +"
    elif bal < -0.01:
        sign = "🔴 -"
    else:
        sign = "⚪ "
    md_lines.append(f"| **{p}** | €{paid[p]:.2f} | €{owes[p]:.2f} | {sign}€{abs(bal):.2f} | {sign}₪{abs(ils(bal)):,.2f} |")

md_lines += [
    "",
    "🟢 = מגיע לו כסף  |  🔴 = חייב כסף",
    "",
    "---",
    "",
    "## 🔁 העברות לביצוע (מינימליות)",
    "",
]

by_recipient = defaultdict(list)
for src, dst, amt in transfers:
    by_recipient[dst].append((src, amt))

for recipient in sorted(by_recipient.keys(), key=lambda r: -sum(a for _, a in by_recipient[r])):
    total = sum(a for _, a in by_recipient[recipient])
    md_lines.append(f'### → {recipient} (יקבל סה"כ €{total:.2f} / ₪{ils(total):,.2f})')
    md_lines.append("")
    for src, amt in by_recipient[recipient]:
        md_lines.append(f"- **{src}** מעביר **€{amt:.2f}** (₪{ils(amt):,.2f})")
    md_lines.append("")

md_lines += [
    "---",
    "",
    "## 👨‍👩‍👧‍👦 חישוב חלופי — משפחת חורי הגדולה כיחידה אחת",
    "",
    "_אדיר משלם בפועל על כל המשפחה (אדיר, הילה, כפיר, עדי, שירה, נועם)_",
    "",
    "### יתרה מאוחדת",
    "",
    "| יחידה | שילם | חייב | יתרה (€) | יתרה (₪) |",
    "|-------|------|------|------|------|",
]

family_order = ["יהודה", "משפחת חורי (אדיר)", "נאוה", "אגם", "יובל", "חנה", "אור", "ירין"]
for p in family_order:
    bal = balances_family[p]
    if bal > 0.01:
        sign = "🟢 +"
    elif bal < -0.01:
        sign = "🔴 -"
    else:
        sign = "⚪ "
    md_lines.append(f"| **{p}** | €{paid_family[p]:.2f} | €{owes_family[p]:.2f} | {sign}€{abs(bal):.2f} | {sign}₪{abs(ils(bal)):,.2f} |")

md_lines += [
    "",
    "### 🔁 העברות לביצוע (משפחה כיחידה)",
    "",
]

by_recipient_f = defaultdict(list)
for src, dst, amt in transfers_family:
    by_recipient_f[dst].append((src, amt))

for recipient in sorted(by_recipient_f.keys(), key=lambda r: -sum(a for _, a in by_recipient_f[r])):
    total = sum(a for _, a in by_recipient_f[recipient])
    md_lines.append(f'### → {recipient} (יקבל סה"כ €{total:.2f} / ₪{ils(total):,.2f})')
    md_lines.append("")
    for src, amt in by_recipient_f[recipient]:
        md_lines.append(f"- **{src}** מעביר **€{amt:.2f}** (₪{ils(amt):,.2f})")
    md_lines.append("")

md_lines += [
    "---",
    "",
    f"_סה\"כ {len(transfers)} העברות (חישוב מלא) / {len(transfers_family)} העברות (משפחה מאוחדת)_",
    "",
    f"_חושב עם €44.625 לזוג ו-€59.50 לבודד באיזיריידר. הפרש €{total_paid - total_owes:.2f} מעיגול._",
    "",
    f"_שער המרה לשקלים: €1 = ₪{EUR_TO_ILS}_",
]

with open(os.path.join(os.path.dirname(__file__), "data", "athens_settlement.md"), "w", encoding="utf-8") as f:
    f.write("\n".join(md_lines))

print("\n[OK] data/athens_settlement.md")

# ═══════════════════════════════════════════════════════════════════════
# Excel
# ═══════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

HEADER_FILL = PatternFill("solid", fgColor="1A6B3C")
ROW_A_FILL = PatternFill("solid", fgColor="D6F0E0")
ROW_B_FILL = PatternFill("solid", fgColor="B2DFC5")
TOTAL_FILL = PatternFill("solid", fgColor="7EC8A0")
GREEN_FILL = PatternFill("solid", fgColor="C6F1D6")
RED_FILL = PatternFill("solid", fgColor="F8D5D5")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FONT = Font(bold=True, color="1A3A2A")

RTL_ALIGN = Alignment(horizontal="right", readingOrder=2)
RTL_WRAP = Alignment(horizontal="right", readingOrder=2, wrap_text=True)

# גיליון 1: יתרות + העברות
ws1 = wb.active
ws1.title = "סיכום ויתרות"
ws1.sheet_view.rightToLeft = True

ws1["A1"] = "💳 סיכום הוצאות אתונה 2026"
ws1["A1"].font = Font(bold=True, size=14, color="1A3A2A")
ws1.merge_cells("A1:E1")
ws1["A1"].alignment = RTL_ALIGN

ws1["A2"] = f'סה"כ הוצאות: €{total_paid:,.2f} (₪{ils(total_paid):,.2f})  |  משתתפים: 13  |  שער: €1=₪{EUR_TO_ILS}'
ws1.merge_cells("A2:E2")
ws1["A2"].alignment = RTL_ALIGN

# כותרות יתרות
row = 4
for col, h in enumerate(["אדם", "שילם (€)", "חייב (€)", "יתרה (€)", "יתרה (₪)"], 1):
    c = ws1.cell(row=row, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = RTL_ALIGN

for i, p in enumerate(PEOPLE):
    bal = balances[p]
    fill = GREEN_FILL if bal > 0.01 else RED_FILL if bal < -0.01 else ROW_A_FILL
    values = [p, round(paid[p], 2), round(owes[p], 2), round(bal, 2), ils(bal)]
    for col, v in enumerate(values, 1):
        c = ws1.cell(row=row + 1 + i, column=col, value=v)
        c.fill = fill
        c.alignment = RTL_ALIGN
        if col in (4, 5) and bal > 0.01:
            c.font = Font(bold=True, color="1A6B3C")
        elif col in (4, 5) and bal < -0.01:
            c.font = Font(bold=True, color="C0392B")
        if isinstance(v, float):
            c.number_format = "#,##0.00"

for col, w in enumerate([15, 12, 12, 14, 14], 1):
    ws1.column_dimensions[get_column_letter(col)].width = w

# העברות
transfers_start = row + 1 + len(PEOPLE) + 2
hc = ws1.cell(row=transfers_start, column=1, value="🔁 העברות לביצוע")
hc.font = Font(bold=True, size=12)
hc.alignment = RTL_ALIGN
ws1.merge_cells(start_row=transfers_start, start_column=1, end_row=transfers_start, end_column=5)

for col, h in enumerate(["מ", "ל", "סכום (€)", "סכום (₪)", ""], 1):
    c = ws1.cell(row=transfers_start + 1, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = RTL_ALIGN

for i, (src, dst, amt) in enumerate(transfers):
    fill = ROW_A_FILL if i % 2 == 0 else ROW_B_FILL
    values = [src, dst, round(amt, 2), ils(amt), ""]
    for col, v in enumerate(values, 1):
        c = ws1.cell(row=transfers_start + 2 + i, column=col, value=v)
        c.fill = fill
        c.alignment = RTL_ALIGN
        if isinstance(v, float):
            c.number_format = "#,##0.00"

# גיליון 2: פירוט הוצאות
ws2 = wb.create_sheet("פירוט הוצאות")
ws2.sheet_view.rightToLeft = True

for col, h in enumerate(["הוצאה", "סכום (€)", "סכום (₪)", "מי שילם", "משתתפים", "חלוקה"], 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = RTL_ALIGN

for i, exp in enumerate(expenses):
    fill = ROW_A_FILL if i % 2 == 0 else ROW_B_FILL
    paid_str = ", ".join(f"{p} €{a:.2f}" for p, a in exp["paid_by"].items())
    if exp.get("custom_split"):
        participants_str = "13 (זוגות+בודדים)"
        split_str = "€44.625 לזוג / €59.50 לבודד"
    elif exp["split_among"]:
        participants_str = ", ".join(exp["split_among"])
        per_person = exp["total"] / len(exp["split_among"])
        split_str = f"€{per_person:.2f} × {len(exp['split_among'])}"
    else:
        participants_str = "—"
        split_str = "—"

    values = [exp["desc"], round(exp["total"], 2), ils(exp["total"]), paid_str, participants_str, split_str]
    for col, v in enumerate(values, 1):
        c = ws2.cell(row=i + 2, column=col, value=v)
        c.fill = fill
        c.alignment = RTL_WRAP
        if isinstance(v, float):
            c.number_format = "#,##0.00"

total_row = len(expenses) + 2
for col in range(1, 7):
    cell = ws2.cell(row=total_row, column=col)
    cell.fill = TOTAL_FILL
    cell.font = TOTAL_FONT
    cell.alignment = RTL_ALIGN

ws2.cell(row=total_row, column=1, value='סה"כ')
ws2.cell(row=total_row, column=2, value=round(total_paid, 2)).number_format = "#,##0.00"
ws2.cell(row=total_row, column=3, value=ils(total_paid)).number_format = "#,##0.00"

for col, w in enumerate([28, 12, 12, 28, 40, 22], 1):
    ws2.column_dimensions[get_column_letter(col)].width = w

# ─── גיליון 3: משפחת חורי כיחידה ────────────────────────────────────────
ws3 = wb.create_sheet("משפחה כיחידה")
ws3.sheet_view.rightToLeft = True

ws3["A1"] = "👨‍👩‍👧‍👦 משפחת חורי הגדולה כיחידה אחת"
ws3["A1"].font = Font(bold=True, size=14, color="1A3A2A")
ws3.merge_cells("A1:E1")
ws3["A1"].alignment = RTL_ALIGN

ws3["A2"] = "אדיר משלם בפועל על: אדיר, הילה, כפיר, עדי, שירה, נועם"
ws3.merge_cells("A2:E2")
ws3["A2"].alignment = RTL_ALIGN

# כותרות יתרות
row = 4
for col, h in enumerate(["יחידה", "שילם (€)", "חייב (€)", "יתרה (€)", "יתרה (₪)"], 1):
    c = ws3.cell(row=row, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = RTL_ALIGN

for i, p in enumerate(family_order):
    bal = balances_family[p]
    fill = GREEN_FILL if bal > 0.01 else RED_FILL if bal < -0.01 else ROW_A_FILL
    values = [p, round(paid_family[p], 2), round(owes_family[p], 2), round(bal, 2), ils(bal)]
    for col, v in enumerate(values, 1):
        c = ws3.cell(row=row + 1 + i, column=col, value=v)
        c.fill = fill
        c.alignment = RTL_ALIGN
        if col in (4, 5) and bal > 0.01:
            c.font = Font(bold=True, color="1A6B3C")
        elif col in (4, 5) and bal < -0.01:
            c.font = Font(bold=True, color="C0392B")
        if isinstance(v, float):
            c.number_format = "#,##0.00"

for col, w in enumerate([24, 12, 12, 14, 14], 1):
    ws3.column_dimensions[get_column_letter(col)].width = w

# העברות משפחה
transfers_start_f = row + 1 + len(family_order) + 2
hc = ws3.cell(row=transfers_start_f, column=1, value="🔁 העברות לביצוע")
hc.font = Font(bold=True, size=12)
hc.alignment = RTL_ALIGN
ws3.merge_cells(start_row=transfers_start_f, start_column=1, end_row=transfers_start_f, end_column=5)

for col, h in enumerate(["מ", "ל", "סכום (€)", "סכום (₪)", ""], 1):
    c = ws3.cell(row=transfers_start_f + 1, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = RTL_ALIGN

for i, (src, dst, amt) in enumerate(transfers_family):
    fill = ROW_A_FILL if i % 2 == 0 else ROW_B_FILL
    values = [src, dst, round(amt, 2), ils(amt), ""]
    for col, v in enumerate(values, 1):
        c = ws3.cell(row=transfers_start_f + 2 + i, column=col, value=v)
        c.fill = fill
        c.alignment = RTL_ALIGN
        if isinstance(v, float):
            c.number_format = "#,##0.00"

# RTL ב-workbook level
wb.views[0].rightToLeft = True

wb.save(os.path.join(os.path.dirname(__file__), "data", "athens_settlement.xlsx"))
print("[OK] data/athens_settlement.xlsx")
